import re
from tkinter import *
import tkinter as tk
import openpyxl
from tkinter import ttk
import fpdf
import openpyxl
import datetime as dt
from time import strftime
from tkinter.messagebox import showinfo
import sys
import os
import webbrowser
os.chdir(sys.path[0])

# V1.2 Змінений розмір стікера гарантії на 40х24, змінений розмір шрифта та висоту рядка
# V1.3 Змінена орієнтація стікерів гарантії та пакування, з портрету на альбомний вигляд


win = tk.Tk()
img = PhotoImage(file="./img/ECOSOFT_LOGO.png")
img1 = img.subsample(2, 2)
labelLogo = Label(win, image=img1, background='#EBEBEB')
labelLogo.grid(row=0, column=0, columnspan=5,
               rowspan=1, padx=1, pady=0, sticky='', ipady=0)


def InitUI():
    photo = tk.PhotoImage(file='img\logo.png')
    win.iconphoto(False, photo)
    win.config(bg='#EBEBEB')
    win.title(
        'Створення PDF файлів для установок, пакування, гарантійного талона')
    win.geometry('740x545+400+100')
    win.resizable(False, False)


def dropDownList(row):
    chosed_SKU = tk.StringVar()
    combo = ttk.Combobox(win, textvariable=chosed_SKU,
                         values=SKU, width=20)
    combo['state'] = 'readonly'
    combo.grid(column=1, row=row, padx=3, pady=3)


def buttonClicSameSKU():
    sku_variable_2.set(sku_variable.get())
    sku_variable_3.set(sku_variable.get())
    sku_variable_4.set(sku_variable.get())
    sku_variable_5.set(sku_variable.get())
    sku_variable_6.set(sku_variable.get())
    sku_variable_7.set(sku_variable.get())
    sku_variable_8.set(sku_variable.get())
    sku_variable_9.set(sku_variable.get())
    sku_variable_10.set(sku_variable.get())
    sku_variable_11.set(sku_variable.get())
    sku_variable_12.set(sku_variable.get())


def buttonClicListCH():
    serial_variable_2.set(int(serial_variable.get())+1)
    serial_variable_3.set(int(serial_variable.get())+2)
    serial_variable_4.set(int(serial_variable.get())+3)
    serial_variable_5.set(int(serial_variable.get())+4)
    serial_variable_6.set(int(serial_variable.get())+5)
    serial_variable_7.set(int(serial_variable.get())+6)
    serial_variable_8.set(int(serial_variable.get())+7)
    serial_variable_9.set(int(serial_variable.get())+8)
    serial_variable_10.set(int(serial_variable.get())+9)
    serial_variable_11.set(int(serial_variable.get())+10)
    serial_variable_12.set(int(serial_variable.get())+11)


def findIndexSKU(whatIndexISearch):

    listSKU = [sheet[i][0].value for i in range(2, sheet.max_row)]
    inxexSKU = listSKU.index(whatIndexISearch)+1


def PrintShield(pdf, rowSKU, rowCH, rowDate, listSKU, pad_x, pad_y):
    if rowSKU.get() == '':
        return
    readedIndex = listSKU.index(rowSKU.get())+2

    model_SKU = str((sheet[readedIndex][0]).value)
    model_name = str((sheet[readedIndex][1]).value)
    serial_number = str(rowCH.get())
    model_voltage = str((sheet[readedIndex][2]).value)
    model_frk = str((sheet[readedIndex][3]).value)
    model_power = str((sheet[readedIndex][4]).value)
    model_current = str((sheet[readedIndex][5]).value)

    pdf.set_auto_page_break(False)
    pdf.set_margins(2 + pad_x, 2 + pad_y, 2)
    # pdf.add_page()

    pdf.add_font('arialB', '', 'fonts/arial_bold.ttf', uni=True)
    pdf.add_font('arial', '', 'fonts/arial.ttf', uni=True)
    pdf.set_font('arialB', '', 7.5)

    pdf.image("img/frame.png", w=70, h=45, x=1+pad_x, y=1+pad_y)

    pdf.set_xy(pad_x+2, pad_y+2)

    pdf.cell(txt='REVERSE', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='СИСТЕМА', w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='OSMOSIS', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='ЗВОРОТНОГО', w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='SYSTEM', w=34, h=2.5, border=0, align='L')
    pdf.cell(txt='ОСМОСУ', w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(w=0, h=0.4, ln=1)

    pdf.set_font('arial', '', 7.2)

    pdf.cell(txt='Model: ' + model_name, w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Модель: ' + model_name, w=34,
             h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='SKU: ' + model_SKU, w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Код: ' + model_SKU, w=34,
             h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='Manufacturing date:', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Дата виготовлення:', w=34,
             h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt=rowDate.get(), w=34, h=2.8, border=0, align='L')
    pdf.cell(txt=rowDate.get(), w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt=f'Serial no.: {serial_number}',
             w=34, h=2.8, border=0, align='L')
    pdf.cell(txt=f'Серійний №: {serial_number}',
             w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='Supply voltage:', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Напруга живлення:', w=34,
             h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt=model_voltage + ' VAC, ' + model_frk +
             ' Hz', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt=model_voltage + ' В, ' + model_frk +
             ' Гц', w=34, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='Power rating: ' + model_power +
             ' kW', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Потужність: ' + model_power + ' кВт',
             w=35, h=2.8, border=0, align='L', ln=1)
    pdf.cell(txt='Ampere rating: ' + model_current +
             ' A', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Споживаний струм: ' + model_current +
             ' А', w=34, h=3, border=0, align='L', ln=1)
    pdf.cell(txt='Manufacturer:', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='Виробник:', w=34, h=2.8, border=0, align='L', ln=1)

    pdf.set_font('arialB', '', 7.5)
    pdf.cell(w=0, h=0.4, ln=1)
    pdf.cell(txt='Ecosoft SPC LTD', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='ТОВ «НВО «ЕКОСОФТ»', w=34,
             h=2.8, border=0, align='L', ln=1)

    pdf.set_font('arial', '', 7.5)
    pdf.cell(txt='QMS ISO 9001:2015', w=34, h=2.8, border=0, align='L')
    pdf.cell(txt='СУЯ ISO 9001:2015', w=34,
             h=2.8, border=0, align='L', ln=1)

    pdf.image("img\CE.png", x=25 + pad_x, y=2.5 + pad_y, h=7)
    pdf.image("img\EAC.png", x=61 + pad_x, y=2.5 + pad_y, h=7)
    pdf.image("img/ukrsepro.png", x=28 + pad_x, y=35.5 + pad_y, h=9)


def PrintVerticalPackingSticker(pdf, rowSKU, rowCH, rowDate, listSKU):
    readedIndex = listSKU.index(rowSKU.get())+2
    num_line = readedIndex

    raw = sheet[num_line][6]
    raw = str(raw.value)
    # print(raw)

    descrUA = (sheet[num_line][7]).value
    # print(descrUA)

    descrCODE = (sheet[num_line][0]).value
    # print(descrCODE)

    pdf.set_auto_page_break(False)
    pdf.set_margins(2, 3, 3)
    pdf.add_page(orientation='P')

    pdf.image("img\ecosoft.png", x=2, y=3, w=65.06396)

    pdf.image("img\EAC.png", x=74, y=35, h=6.25615)

    pdf.image("img\CE.png", x=72, y=4.5, h=6.881765)

    pdf.image("img/ukrsepro.png", x=73, y=18, h=11.26107)

    pdf.image(f"img\{rowSKU.get()}.png", x=2, y=55, h=33.032472)

    pdf.add_font('arial', '', 'fonts/arial.ttf', uni=True)
    pdf.add_font('arialB', '', 'fonts/arial_black.ttf', uni=True)

    pdf.set_font('arial', '', 6.5)

    pdf.set_y(18.4), pdf.set_x(9)
    pdf.cell(txt='Storage temperature from +5 to +40 °C (41 to 104 °F)',
             w=60, h=2.5, border=0)
    pdf.set_y(21.2), pdf.set_x(9)
    pdf.cell(txt='Manufacturer: Ecosoft SPC LTD,',
             w=60, h=2.5, border=0)
    pdf.set_y(24.0), pdf.set_x(9)
    pdf.cell(txt='1Ї, Pokrovska str. Irpin, Kyiv Oblast, 08200, Ukraine',
             w=60, h=2.5, border=0)

    pdf.set_y(27.5), pdf.set_x(9)
    pdf.cell(txt='Зберігати за температури від +5 до +40 °C',
             w=60, h=2.5, border=0)
    pdf.set_y(30.3), pdf.set_x(9)
    pdf.cell(txt='Виробник: ТОВ «НВО «ЕКОСОФТ», вул. Покровська, 1Ї',
             w=60, h=2.5, border=0)
    pdf.set_y(33.0), pdf.set_x(9)
    pdf.cell(txt='м. Ірпінь, Київська область, 08200, Україна',
             w=60, h=2.5, border=0)
    pdf.set_y(38.0), pdf.set_x(9)
    pdf.cell(txt='ТУУ 28.2-31749798-005:2013',
             w=60, h=2.5, border=0)
    pdf.set_y(41.0), pdf.set_x(9)
    pdf.cell(txt='Ecosoft — BWT Group',
             w=60, h=2.5, border=0)

    pdf.set_font('arial', '', 9)

    pdf.set_y(45), pdf.set_x(2)
    pdf.cell(txt='Manufacture date /', w=0, h=5, border=0, ln=2)
    pdf.cell(txt='Дата виготовлення:', w=0, h=5, border=0, ln=2)

    pdf.set_y(47), pdf.set_x(35)
    current_date = rowDate.get()
    pdf.cell(txt=current_date, w=0, h=5, border=0,
             ln=2, align='L')

    pdf.set_y(47), pdf.set_x(60)
    serial_number = f"CH-{rowCH.get()}"
    pdf.cell(txt=serial_number, w=0, h=5, border=0, ln=2)

    pdf.set_y(55), pdf.set_x(2)
    pdf.cell(w=0, h=33, border=0, ln=1)

    pdf.add_font('Dejavu', '', 'fonts/arialn.ttf', uni=True)
    pdf.set_font('Dejavu', '', 6.1)

    pdf.cell(w=0, h=4.0, txt=raw, border=0,
             ln=1)
    pdf.cell(w=0, h=4.0, txt=descrUA, border=0,
             ln=1)

    pdf.set_font('arialB', '', size=19)
    pdf.cell(w=0, h=6.0, txt=descrCODE, border=0,
             ln=1)


def PrintHorizontalPackingSticker(pdf, rowSKU, rowCH, rowDate, listSKU):
    readedIndex = listSKU.index(rowSKU.get())+2
    num_line = readedIndex

    raw = sheet[num_line][6]
    raw = str(raw.value)
    # print(raw)

    descrUA = (sheet[num_line][7]).value
    # print(descrUA)

    descrCODE = (sheet[num_line][0]).value
    # print(descrCODE)

    pdf.set_auto_page_break(False)
    pdf.set_margins(2, 3, 3)
    pdf.add_page(orientation='L')

    pdf.image("img\ecosoft.png", x=2, y=3, w=65.06396)

    pdf.image("img\EAC.png", x=34, y=48, h=6.25615)

    pdf.image("img\CE.png", x=10, y=48, h=6.881765)

    pdf.image("img/ukrsepro.png", x=56, y=48, h=11.26107)

    pdf.image(f"img\{rowSKU.get()}.png", x=73, y=29, w=28.152675)

    pdf.add_font('arial', '', 'fonts/arial.ttf', uni=True)
    pdf.add_font('arialB', '', 'fonts/arial_black.ttf', uni=True)

    pdf.set_font('arial', '', 6.5)

    pdf.set_y(18.4), pdf.set_x(9)
    pdf.cell(txt='Storage temperature from +5 to +40 °C (41 to 104 °F)',
             w=60, h=2.5, border=0)
    pdf.set_y(21.2), pdf.set_x(9)
    pdf.cell(txt='Manufacturer: Ecosoft SPC LTD,',
             w=60, h=2.5, border=0)
    pdf.set_y(24.0), pdf.set_x(9)
    pdf.cell(txt='1Ї, Pokrovska str. Irpin, Kyiv Oblast, 08200, Ukraine',
             w=60, h=2.5, border=0)

    pdf.set_y(27.5), pdf.set_x(9)
    pdf.cell(txt='Зберігати за температури від +5 до +40 °C',
             w=60, h=2.5, border=0)
    pdf.set_y(30.3), pdf.set_x(9)
    pdf.cell(txt='Виробник: ТОВ «НВО «ЕКОСОФТ», вул. Покровська, 1Ї',
             w=60, h=2.5, border=0)
    pdf.set_y(33.0), pdf.set_x(9)
    pdf.cell(txt='м. Ірпінь, Київська область, 08200, Україна',
             w=60, h=2.5, border=0)
    pdf.set_y(38.0), pdf.set_x(9)
    pdf.cell(txt='ТУУ 28.2-31749798-005:2013',
             w=60, h=2.5, border=0)
    pdf.set_y(41.0), pdf.set_x(9)
    pdf.cell(txt='Ecosoft — BWT Group',
             w=60, h=2.5, border=0)

    pdf.set_font('arial', '', 9)

    pdf.set_y(7), pdf.set_x(70)
    pdf.cell(txt='Manufacture date /', w=0, h=5, border=0, ln=2)
    pdf.cell(txt='Дата виготовлення:', w=0, h=5, border=0, ln=2)

    current_date = rowDate.get()

    pdf.cell(txt=current_date, w=0, h=5, border=0,
             ln=2, align='L')

    serial_number = f"CH-{rowCH.get()}"

    pdf.cell(txt=serial_number, w=0, h=5, border=0, ln=2)
    pdf.cell(w=0, h=32, border=0, ln=1)

    pdf.add_font('Dejavu', '', 'fonts/arialn.ttf', uni=True)
    pdf.set_font('Dejavu', '', 6.1)

    pdf.cell(w=68, h=6.0, txt=raw, border=0,
             ln=1)
    pdf.cell(w=0, h=10.0, txt=descrUA, border=0,
             ln=1)

    pdf.set_font('arialB', '', size=19)
    pdf.write(txt=descrCODE)


def PrintPackingSticker(pdf, rowSKU, rowCH, rowDate, listSKU):
    if rowSKU.get() == '':
        return
    if (rowSKU.get() == 'MO12XLWE0UN'
        or rowSKU.get() == 'MO16XLWE0UN'
        or rowSKU.get() == 'MO20XLWE0UN'
        or rowSKU.get() == 'MO24XLWE0UN'
        or rowSKU.get() == 'MO30XLWE0UN'
        or rowSKU.get() == 'MO36XLWE0UN'
            or rowSKU.get() == 'MO48XLWE0UN'):
        PrintVerticalPackingSticker(pdf, rowSKU, rowCH, rowDate, listSKU)
    else:
        PrintHorizontalPackingSticker(pdf, rowSKU, rowCH, rowDate, listSKU)


def PrintGuaranteeSticker(pdf, rowSKU, rowCH, rowDate):
    if rowSKU.get() == '':
        return
    model_SKU = rowSKU.get()
    serial_number = rowCH

    pdf.set_auto_page_break(False)
    pdf.set_margins(2, 2, 2)
    pdf.add_page(orientation='L')
    pdf.add_font('arial', '', 'fonts/arial.ttf', uni=True)
    pdf.set_font('arial', '', 12)
    pdf.cell(txt=model_SKU, align='C', w=0, h=7, border=0, ln=2)
    pdf.cell(txt=f'CH-{serial_number.get()}',
             align='C', w=0, h=7, border=0, ln=2)
    pdf.cell(txt=f'Date: {rowDate.get()}', w=0, h=7, border=0, align='C')

# Створення 12 стікерів


def Shields_12pcs():
    move_x_0 = 31.5
    move_x_1 = 72
    move_y_0 = 8
    move_y_1 = 46.5
    # pdf = fpdf.FPDF('P', 'mm', (72, 47))
    pdf = fpdf.FPDF('P', 'mm', (210, 297))
    pdf.add_page()
    PrintShield(pdf, sku_variable, serial_variable,
                date_variable, SKU, move_x_0, move_y_0)

    PrintShield(pdf, sku_variable_2, serial_variable_2,
                date_variable_2, SKU, move_x_0 + move_x_1, move_y_0)

    PrintShield(pdf, sku_variable_3, serial_variable_3,
                date_variable_3, SKU, move_x_0, move_y_0 + move_y_1)

    PrintShield(pdf, sku_variable_4, serial_variable_4,
                date_variable_4, SKU, move_x_0 + move_x_1, move_y_0 + move_y_1)

    PrintShield(pdf, sku_variable_5, serial_variable_5,
                date_variable_5, SKU, move_x_0, move_y_0 + move_y_1*2)

    PrintShield(pdf, sku_variable_6, serial_variable_6,
                date_variable_6, SKU, move_x_0 + move_x_1, move_y_0 + move_y_1*2)

    PrintShield(pdf, sku_variable_7, serial_variable_7,
                date_variable_7, SKU, move_x_0, move_y_0 + move_y_1*3)

    PrintShield(pdf, sku_variable_8, serial_variable_8,
                date_variable_8, SKU, move_x_0 + move_x_1, move_y_0 + move_y_1*3)

    PrintShield(pdf, sku_variable_9, serial_variable_9,
                date_variable_9, SKU, move_x_0, move_y_0 + move_y_1*4)

    PrintShield(pdf, sku_variable_10, serial_variable_10,
                date_variable_10, SKU, move_x_0 + move_x_1, move_y_0 + move_y_1*4)

    PrintShield(pdf, sku_variable_11, serial_variable_11,
                date_variable_11, SKU, move_x_0, move_y_0 + move_y_1*5)

    PrintShield(pdf, sku_variable_12, serial_variable_12,
                date_variable_12, SKU, move_x_0 + move_x_1, move_y_0 + move_y_1*5)

    pdf.output("Шильди.pdf")


def GuaranteeStikers_12pcs():

    pdf = fpdf.FPDF('P', 'mm', (24, 40))
    PrintGuaranteeSticker(
        pdf, sku_variable, serial_variable, date_variable)
    PrintGuaranteeSticker(pdf, sku_variable_2,
                          serial_variable_2, date_variable_2)
    PrintGuaranteeSticker(pdf, sku_variable_3,
                          serial_variable_3, date_variable_3)
    PrintGuaranteeSticker(pdf, sku_variable_4,
                          serial_variable_4, date_variable_4)
    PrintGuaranteeSticker(pdf, sku_variable_5,
                          serial_variable_5, date_variable_5)
    PrintGuaranteeSticker(pdf, sku_variable_6,
                          serial_variable_6, date_variable_6)
    PrintGuaranteeSticker(pdf, sku_variable_7,
                          serial_variable_7, date_variable_7)
    PrintGuaranteeSticker(pdf, sku_variable_8,
                          serial_variable_8, date_variable_8)
    PrintGuaranteeSticker(pdf, sku_variable_9,
                          serial_variable_9, date_variable_9)
    PrintGuaranteeSticker(pdf, sku_variable_10,
                          serial_variable_10, date_variable_10)
    PrintGuaranteeSticker(pdf, sku_variable_11,
                          serial_variable_11, date_variable_11)
    PrintGuaranteeSticker(pdf, sku_variable_12,
                          serial_variable_12, date_variable_12)
    pdf.output("Стікери гарантії.pdf")

    return


def PackingStikers_12pcs():

    pdf = fpdf.FPDF('P', 'mm', (86, 104))
    PrintPackingSticker(pdf, sku_variable,
                        serial_variable, date_variable, SKU)
    PrintPackingSticker(pdf, sku_variable_2,
                        serial_variable_2, date_variable_2, SKU)
    PrintPackingSticker(pdf, sku_variable_3,
                        serial_variable_3, date_variable_3, SKU)
    PrintPackingSticker(pdf, sku_variable_4,
                        serial_variable_4, date_variable_4, SKU)
    PrintPackingSticker(pdf, sku_variable_5,
                        serial_variable_5, date_variable_5, SKU)
    PrintPackingSticker(pdf, sku_variable_6,
                        serial_variable_6, date_variable_6, SKU)
    PrintPackingSticker(pdf, sku_variable_7,
                        serial_variable_7, date_variable_7, SKU)
    PrintPackingSticker(pdf, sku_variable_8,
                        serial_variable_8, date_variable_8, SKU)
    PrintPackingSticker(pdf, sku_variable_9,
                        serial_variable_9, date_variable_9, SKU)
    PrintPackingSticker(pdf, sku_variable_10,
                        serial_variable_10, date_variable_10, SKU)
    PrintPackingSticker(pdf, sku_variable_11,
                        serial_variable_11, date_variable_11, SKU)
    PrintPackingSticker(pdf, sku_variable_12,
                        serial_variable_12, date_variable_12, SKU)
    pdf.output("Стікери упаковки.pdf")


def PrintShields_12pcs():
    if sku_variable.get() == "":
        showinfo(message="Оберіть код товару.")
        return
    if serial_variable.get() == "":
        showinfo(message="Введіть серійний номер.")
        return
    Shields_12pcs()
    showinfo(message='Шильди збережено!')


def PrintGuaranteeStikers_12pcs():
    if sku_variable.get() == "":
        showinfo(message="Оберіть код товару.")
        return
    if serial_variable.get() == "":
        showinfo(message="Введіть серійний номер.")
        return
    GuaranteeStikers_12pcs()
    showinfo(message='Стікери гарантії збережено!')


def PrintPackingStikers_12pcs():
    if sku_variable.get() == "":
        showinfo(message="Оберіть код товару.")
        return
    if serial_variable.get() == "":
        showinfo(message="Введіть серійний номер.")
        return
    PackingStikers_12pcs()
    showinfo(message='Стікери упаковки збережено!')


def SaveAllStickers():
    if sku_variable.get() == "":
        showinfo(message="Оберіть код товару.")
        return
    if serial_variable.get() == "":
        showinfo(message="Введіть серійний номер.")
        return
    Shields_12pcs()
    GuaranteeStikers_12pcs()
    PackingStikers_12pcs()
    showinfo(message="Шильди, стікери гарантії та пакування збережені.")


def openAuthorLink(url):
    webbrowser.open_new(url)


# Час
dateCurrent = dt.datetime.now()
dateLabel = Label(win, font="arial, 20")
time_string = strftime('%d-%m-%Y')

# Початок програми

InitUI()

# Створення змінних для першого рядку
sku_variable = tk.StringVar(win)
date_variable = tk.StringVar(win)
serial_variable = tk.StringVar(win)

# Створення змінних для другого рядку
sku_variable_2 = tk.StringVar(win)
date_variable_2 = tk.StringVar(win)
serial_variable_2 = tk.StringVar(win)
# Створення змінних для третього рядку
sku_variable_3 = tk.StringVar(win)
date_variable_3 = tk.StringVar(win)
serial_variable_3 = tk.StringVar(win)
# Створення змінних для четвертого рядку
sku_variable_4 = tk.StringVar(win)
date_variable_4 = tk.StringVar(win)
serial_variable_4 = tk.StringVar(win)
# Створення змінних для п'ятого рядку
sku_variable_5 = tk.StringVar(win)
date_variable_5 = tk.StringVar(win)
serial_variable_5 = tk.StringVar(win)
# Створення змінних для шостого рядку
sku_variable_6 = tk.StringVar(win)
date_variable_6 = tk.StringVar(win)
serial_variable_6 = tk.StringVar(win)
# Створення змінних для сьомого рядку
sku_variable_7 = tk.StringVar(win)
date_variable_7 = tk.StringVar(win)
serial_variable_7 = tk.StringVar(win)
# Створення змінних для восьмого рядку
sku_variable_8 = tk.StringVar(win)
date_variable_8 = tk.StringVar(win)
serial_variable_8 = tk.StringVar(win)

# Створення змінних для дев'ятого рядку
sku_variable_9 = tk.StringVar(win)
date_variable_9 = tk.StringVar(win)
serial_variable_9 = tk.StringVar(win)

# Створення змінних для десятого рядку
sku_variable_10 = tk.StringVar(win)
date_variable_10 = tk.StringVar(win)
serial_variable_10 = tk.StringVar(win)

# Створення змінних для одинадцятого рядку
sku_variable_11 = tk.StringVar(win)
date_variable_11 = tk.StringVar(win)
serial_variable_11 = tk.StringVar(win)

# Створення змінних для дванадцятого рядку
sku_variable_12 = tk.StringVar(win)
date_variable_12 = tk.StringVar(win)
serial_variable_12 = tk.StringVar(win)

book = openpyxl.open("data\Sheet1.xlsx", read_only=True)
sheet = book["Sheet1"]
SKU = [sheet[x][0].value for x in range(2, sheet.max_row)]
# print(SKU)
# print(len(SKU))

#   Назви колонок
Label(win, text='Код продукту', bg='#EBEBEB').grid(row=1, column=1)
Label(win, text='Дата виготовлення', bg='#EBEBEB',
      padx=20). grid(row=1, column=2)
Label(win, text='', bg='#EBEBEB', padx=20).grid(
    row=1, column=3, columnspan=1)
Label(win, text='Серійний номер', bg='#EBEBEB').grid(row=1, column=4)

#   Створення стовпчика нумерації
for i in range(1, 13):
    Label(win, text=f'{i}', bg='#EBEBEB', padx=10).grid(row=i+1, column=0)

# Створення комірок SKU
chosed_SKU = tk.StringVar()
combo = ttk.Combobox(win, textvariable=sku_variable, values=SKU, width=20)
combo['state'] = 'readonly'
combo.grid(column=1, row=2, padx=3, pady=3)

combo2 = ttk.Combobox(
    win, textvariable=sku_variable_2, values=SKU, width=20)
combo2['state'] = 'readonly'
combo2.grid(column=1, row=3, padx=3, pady=3)

combo3 = ttk.Combobox(
    win, textvariable=sku_variable_3, values=SKU, width=20)
combo3['state'] = 'readonly'
combo3.grid(column=1, row=4, padx=3, pady=3)

combo4 = ttk.Combobox(
    win, textvariable=sku_variable_4, values=SKU, width=20)
combo4['state'] = 'readonly'
combo4.grid(column=1, row=5, padx=3, pady=3)

combo5 = ttk.Combobox(
    win, textvariable=sku_variable_5, values=SKU, width=20)
combo5['state'] = 'readonly'
combo5.grid(column=1, row=6, padx=3, pady=3)

combo6 = ttk.Combobox(
    win, textvariable=sku_variable_6, values=SKU, width=20)
combo6['state'] = 'readonly'
combo6.grid(column=1, row=7, padx=3, pady=3)

combo7 = ttk.Combobox(
    win, textvariable=sku_variable_7, values=SKU, width=20)
combo7['state'] = 'readonly'
combo7.grid(column=1, row=8, padx=3, pady=3)

combo8 = ttk.Combobox(
    win, textvariable=sku_variable_8, values=SKU, width=20)
combo8['state'] = 'readonly'
combo8.grid(column=1, row=9, padx=3, pady=3)

combo9 = ttk.Combobox(
    win, textvariable=sku_variable_9, values=SKU, width=20)
combo9['state'] = 'readonly'
combo9.grid(column=1, row=10, padx=3, pady=3)

combo10 = ttk.Combobox(
    win, textvariable=sku_variable_10, values=SKU, width=20)
combo10['state'] = 'readonly'
combo10.grid(column=1, row=11, padx=3, pady=3)

combo11 = ttk.Combobox(
    win, textvariable=sku_variable_11, values=SKU, width=20)
combo11['state'] = 'readonly'
combo11.grid(column=1, row=12, padx=3, pady=3)

combo12 = ttk.Combobox(
    win, textvariable=sku_variable_12, values=SKU, width=20)
combo12['state'] = 'readonly'
combo12.grid(column=1, row=13, padx=3, pady=3)

# Комірки дати
entryDate = Entry(win, textvariable=date_variable, justify='center')
entryDate.insert(0, time_string)
entryDate.grid(row=2, column=2, pady=3)

entryDate2 = Entry(win, textvariable=date_variable_2, justify='center')
entryDate2.insert(0, time_string)
entryDate2.grid(row=3, column=2, pady=3)

entryDate3 = Entry(win, textvariable=date_variable_3, justify='center')
entryDate3.insert(0, time_string)
entryDate3.grid(row=4, column=2, pady=3)

entryDate4 = Entry(win, textvariable=date_variable_4, justify='center')
entryDate4.insert(0, time_string)
entryDate4.grid(row=5, column=2, pady=3)

entryDate5 = Entry(win, textvariable=date_variable_5, justify='center')
entryDate5.insert(0, time_string)
entryDate5.grid(row=6, column=2, pady=3)

entryDate6 = Entry(win, textvariable=date_variable_6, justify='center')
entryDate6.insert(0, time_string)
entryDate6.grid(row=7, column=2, pady=3)

entryDate7 = Entry(win, textvariable=date_variable_7, justify='center')
entryDate7.insert(0, time_string)
entryDate7.grid(row=8, column=2, pady=3)

entryDate8 = Entry(win, textvariable=date_variable_8, justify='center')
entryDate8.insert(0, time_string)
entryDate8.grid(row=9, column=2, pady=3)

entryDate9 = Entry(win, textvariable=date_variable_9, justify='center')
entryDate9.insert(0, time_string)
entryDate9.grid(row=10, column=2, pady=3)

entryDate10 = Entry(win, textvariable=date_variable_10, justify='center')
entryDate10.insert(0, time_string)
entryDate10.grid(row=11, column=2, pady=3)

entryDate11 = Entry(win, textvariable=date_variable_11, justify='center')
entryDate11.insert(0, time_string)
entryDate11.grid(row=12, column=2, pady=3)

entryDate12 = Entry(win, textvariable=date_variable_12, justify='center')
entryDate12.insert(0, time_string)
entryDate12.grid(row=13, column=2, pady=3)

#   Створення стовпчика CH-
for i in range(2, 14):
    Label(win, text='СH-', bg='#EBEBEB').grid(row=i, column=3, sticky='E')

# Комірка серійного номеру
Entry(win, textvariable=serial_variable).grid(row=2, column=4, pady=3)
Entry(win, textvariable=serial_variable_2).grid(row=3, column=4, pady=3)
Entry(win, textvariable=serial_variable_3).grid(row=4, column=4, pady=3)
Entry(win, textvariable=serial_variable_4).grid(row=5, column=4, pady=3)
Entry(win, textvariable=serial_variable_5).grid(row=6, column=4, pady=3)
Entry(win, textvariable=serial_variable_6).grid(row=7, column=4, pady=3)
Entry(win, textvariable=serial_variable_7).grid(row=8, column=4, pady=3)
Entry(win, textvariable=serial_variable_8).grid(row=9, column=4, pady=3)
Entry(win, textvariable=serial_variable_9).grid(row=10, column=4, pady=3)
Entry(win, textvariable=serial_variable_10).grid(row=11, column=4, pady=3)
Entry(win, textvariable=serial_variable_11).grid(row=12, column=4, pady=3)
Entry(win, textvariable=serial_variable_12).grid(row=13, column=4, pady=3)

#   Створення кнопок
Button(win, text='SKU однаковий', command=buttonClicSameSKU, width=18, highlightcolor='#ff0000').grid(
    row=14, column=1, ipadx=2, ipady=0, padx=0, pady=5)
Button(win, text='СН по порядку', command=buttonClicListCH, width=16).grid(
    row=14, column=4, ipadx=2, ipady=0, padx=0, pady=0)

Button(win, text='Зберегти шильди', command=PrintShields_12pcs).grid(
    row=2, column=5, ipadx=27, ipady=0, padx=35, pady=0)
Button(win, text='Зберегти стікери гарантії ', command=PrintGuaranteeStikers_12pcs
       ).grid(row=3, column=5, ipadx=5, ipady=0, padx=35, pady=0)
Button(win, text='Зберегти стікери упаковки', command=PrintPackingStikers_12pcs
       ).grid(row=4, column=5, ipadx=2, ipady=0, padx=35, pady=0)

Button(win, text='Зберегти все', command=SaveAllStickers).grid(
    row=6, column=5, ipadx=40, ipady=0, padx=35, pady=0)

#   Напис "розробник"
authorLink = Label(
    win, text='maksym.protsak@gmail.com V1.3', bg='#EBEBEB', fg="blue", cursor="hand2")
authorLink.grid(row=14, column=5, rowspan=5, sticky='ES')
authorLink.bind(
    "<Button-1>", lambda e: openAuthorLink("https://tangerine-youtiao-a51230.netlify.app/"))

#   Запуск програми
win.mainloop()
